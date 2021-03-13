from PIL import Image
from matplotlib import pyplot as plt
from matplotlib import image as img
import os.path
import glob
import xlrd
import openpyxl

data_set = "2021MCMProblemC_DataSet.xlsx"
images_by_global_id = "2021MCM_ProblemC_ Images_by_GlobalID.xlsx"
out_dir = "dataset\\train\\N"


def convert_jpg(jpg_file, out_dir, width=128, height=128):
    b_img = Image.open(jpg_file)
    try:
        new_img = b_img.resize((width, height), Image.BILINEAR)
        new_img.save(os.path.join(out_dir, os.path.basename(jpg_file)))
    except Exception as e:
        print(e)


def data_set_image():
    data = xlrd.open_workbook(data_set)
    data2 = xlrd.open_workbook(images_by_global_id)
    sheet = data.sheets()[0]
    sheet2 = data2.sheets()[0]
    date_rows_num = sheet.nrows
    date2_rows_num = sheet2.nrows
    positive_list = []
    for i in range(1, date_rows_num):
        if sheet.cell(i, 3).value == "Positive ID":

            for j in range(1, date2_rows_num):
                if sheet2.cell(j, 1).value == sheet.cell(i, 0).value:
                    print(sheet2.cell(j, 0))
                    positive_list.append(sheet2.cell(j, 0).value)

    plt.figure()
    for i in range(14):
        plt.subplot(2, 7, i + 1)
        convert_jpg("2021MCM_ProblemC_Files/" + positive_list[i], out_dir)

        a_img = img.imread("img/" + positive_list[i])
        plt.imshow(a_img)
        plt.axis("off")

    plt.suptitle("Asian Giant Hornets Images From the Dataset")
    plt.show()


def web_images():
    plt.figure()
    num = 1
    for image in glob.glob("T1/*.jpg"):
        plt.subplot(7, 8, num)
        num = num + 1
        convert_jpg(image, out_dir + "_web")
        a_img = img.imread("img_web\\" + image[3:])
        plt.imshow(a_img)
        plt.axis("off")

    plt.suptitle("Asian Giant Hornets Images From Web")
    plt.show()


def train_y_images():
    for image in glob.glob("dataset/train/Y/*.jpg"):
        convert_jpg(image, out_dir)


def train_n_images():
    data = xlrd.open_workbook(data_set)
    data2 = xlrd.open_workbook(images_by_global_id)
    sheet = data.sheets()[0]
    sheet2 = data2.sheets()[0]
    date_rows_num = sheet.nrows
    date2_rows_num = sheet2.nrows
    negative_list = []
    for i in range(1, date_rows_num):
        if sheet.cell(i, 3).value == "Negative ID":

            for j in range(1, date2_rows_num):
                if sheet2.cell(j, 1).value == sheet.cell(i, 0).value and sheet2.cell(j, 2).value == "image/jpg":
                    negative_list.append(sheet2.cell(j, 0).value)

    for i in range(len(negative_list)):
        convert_jpg("2021MCM_ProblemC_Files/" + negative_list[i], out_dir)


def unverified_images():
    data = xlrd.open_workbook(data_set)
    data2 = xlrd.open_workbook(images_by_global_id)
    sheet = data.sheets()[0]
    sheet2 = data2.sheets()[0]
    date_rows_num = sheet.nrows
    date2_rows_num = sheet2.nrows
    unverified_list = []
    for i in range(1, date_rows_num):
        if sheet.cell(i, 3).value == "Unverified" or sheet.cell(i, 3).value == "Unprocessed":

            for j in range(1, date2_rows_num):
                if sheet2.cell(j, 1).value == sheet.cell(i, 0).value and sheet2.cell(j, 2).value == "image/jpg":
                    unverified_list.append(sheet2.cell(j, 0).value)

    this_out_dir = "dataset\\unverified"
    for i in range(len(unverified_list)):
        convert_jpg("2021MCM_ProblemC_Files/" + unverified_list[i], this_out_dir)


def rename(old_path, new_path, pre):
    i = 0
    file_list = os.listdir(old_path)
    for files in file_list:
        old_dir_path = os.path.join(old_path, files)
        file_type = os.path.splitext(files)[1]
        new_dir_path = os.path.join(new_path, pre + str(i) + file_type)
        os.rename(old_dir_path, new_dir_path)
        i = i + 1


def label_generate(a_dir, label):
    files = os.listdir(a_dir)
    with open(a_dir + 'val.txt', 'a+') as f:
        for file in files:
            filename = os.path.split(file)[0]
            filetype = os.path.split(file)[1]
            if filetype == '.txt':
                continue
            name = '/all' + '/' + file + ' ' + str(int(label)) + '\n'
            f.write(name)
    print("finished!")


def remark():
    data = xlrd.open_workbook(data_set)
    data2 = xlrd.open_workbook(images_by_global_id)
    sheet = data.sheets()[0]
    sheet2 = data2.sheets()[0]
    date_rows_num = sheet.nrows
    date2_rows_num = sheet2.nrows

    workbook = openpyxl.load_workbook(data_set)
    worksheet = workbook.worksheets[0]
    for i in range(1, date_rows_num):
        if sheet.cell(i, 3).value == "Unverified" or sheet.cell(i, 3).value == "Unprocessed":
            for j in range(1, date2_rows_num):
                if sheet2.cell(j, 1).value == sheet.cell(i, 0).value and sheet2.cell(j, 2).value == "image/jpg":
                    worksheet.cell(i+1, 4, "New Negative ID")
                    break

    workbook.save(data_set)


if __name__ == '__main__':
    remark()
